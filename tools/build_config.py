#!/usr/bin/env python3
"""Build game JSON config from the designer-facing Excel workbook.

Normal workflow:
    python tools/build_config.py

Recovery/bootstrap workflow:
    python tools/build_config.py --from-json
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / "config"
WORKBOOK_PATH = CONFIG_DIR / "GameConfig.xlsx"

UNITS_JSON = CONFIG_DIR / "units.json"
CARDS_JSON = CONFIG_DIR / "cards.json"
LEVELS_JSON = CONFIG_DIR / "levels.json"

UNIT_HEADERS = [
    "unit_id",
    "name",
    "hp",
    "damage",
    "attack_interval_s",
    "move_speed_tiles_per_s",
    "attack_range_tiles",
    "aggro_radius_tiles",
    "body_radius_tiles",
    "unit_type",
    "attack_targets",
    "death_spawn_unit",
    "death_spawn_count",
    "notes",
]
CARD_HEADERS = ["card_id", "name", "elixir_cost", "category", "enabled", "notes"]
SKILL_HEADERS = [
    "card_id",
    "order",
    "skill_type",
    "unit_id",
    "count",
    "damage",
    "radius",
    "target",
    "notes",
]
LEVEL_HEADERS = [
    "level_id",
    "name",
    "elixir_regen_rate",
    "elixir_max",
    "match_duration",
    "ai_difficulty",
    "king_tower_hp",
    "princess_tower_hp",
    "notes",
]
DECK_HEADERS = [
    "level_id",
    "side",
    "slot_1",
    "slot_2",
    "slot_3",
    "slot_4",
    "slot_5",
    "slot_6",
    "slot_7",
    "slot_8",
    "notes",
]

SKILL_TYPES = ["spawn_unit", "direct_damage", "aoe_damage", "aoe_heal"]
UNIT_TYPES = ["ground", "air"]
ATTACK_TARGETS = ["ground", "air", "both"]   # 该单位能攻击的目标类型（V3-2 对空克制）
TARGETS = ["first_enemy_in_lane"]
SIDES = ["player", "ai"]
DIFFICULTIES = ["rookie", "easy", "normal", "hard", "extreme"]


class ConfigError(RuntimeError):
    pass


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any, field: str) -> int | float:
    if value is None or value == "":
        raise ConfigError(f"missing number: {field}")
    if isinstance(value, bool):
        raise ConfigError(f"{field} must be a number, got boolean")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ConfigError(f"{field} must be a number, got {value!r}") from exc
    if not math.isfinite(number):
        raise ConfigError(f"{field} must be finite")
    if number.is_integer():
        return int(number)
    return number


def _number_float(value: Any, field: str) -> float:
    return float(_number(value, field))


def _bool(value: Any) -> bool:
    if value is None or value == "":
        return True
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() not in {"0", "false", "no", "n", "off"}


def _read_rows(wb, sheet_name: str, headers: list[str]) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        raise ConfigError(f"missing sheet: {sheet_name}")
    ws = wb[sheet_name]
    actual = [ws.cell(row=1, column=i + 1).value for i in range(len(headers))]
    if actual != headers:
        raise ConfigError(f"{sheet_name} headers mismatch: expected {headers}, got {actual}")

    rows: list[dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        values = {headers[i]: ws.cell(row=row_index, column=i + 1).value for i in range(len(headers))}
        if all(_is_blank(v) for v in values.values()):
            continue
        values["_row"] = row_index
        rows.append(values)
    return rows


def _require_id(row: dict[str, Any], field: str, sheet: str) -> str:
    value = _text(row.get(field))
    if not value:
        raise ConfigError(f"{sheet}!row {row['_row']} missing {field}")
    return value


def _write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def build_json_from_workbook(workbook_path: Path = WORKBOOK_PATH) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True)

    unit_rows = _read_rows(wb, "Units", UNIT_HEADERS)
    card_rows = _read_rows(wb, "Cards", CARD_HEADERS)
    skill_rows = _read_rows(wb, "CardSkills", SKILL_HEADERS)
    level_rows = _read_rows(wb, "Levels", LEVEL_HEADERS)
    deck_rows = _read_rows(wb, "Decks", DECK_HEADERS)

    units: dict[str, Any] = {}
    for row in unit_rows:
        unit_id = _require_id(row, "unit_id", "Units")
        if unit_id in units:
            raise ConfigError(f"duplicate unit_id: {unit_id}")
        unit_type = _text(row.get("unit_type"))
        if unit_type not in UNIT_TYPES:
            raise ConfigError(f"unit {unit_id} unit_type must be one of {UNIT_TYPES}")
        attack_targets = _text(row.get("attack_targets")) or "ground"
        if attack_targets not in ATTACK_TARGETS:
            raise ConfigError(f"unit {unit_id} attack_targets must be one of {ATTACK_TARGETS}")
        # V3：attack_range / move_speed 量纲改为 tile（attack_range ≥0，无上限；非 lane 比例）。
        attack_range = _number_float(row.get("attack_range_tiles"), f"unit {unit_id}.attack_range_tiles")
        if attack_range < 0.0:
            raise ConfigError(f"unit {unit_id}.attack_range_tiles must be >= 0")
        units[unit_id] = {
            "name": _text(row.get("name")),
            "hp": _number(row.get("hp"), f"unit {unit_id}.hp"),
            "damage": _number(row.get("damage"), f"unit {unit_id}.damage"),
            "attack_speed": _number_float(row.get("attack_interval_s"), f"unit {unit_id}.attack_interval_s"),
            "move_speed": _number_float(row.get("move_speed_tiles_per_s"), f"unit {unit_id}.move_speed_tiles_per_s"),
            "attack_range": attack_range,
            "aggro_radius": _number_float(row.get("aggro_radius_tiles"), f"unit {unit_id}.aggro_radius_tiles"),
            "body_radius": _number_float(row.get("body_radius_tiles"), f"unit {unit_id}.body_radius_tiles"),
            "target_type": unit_type,
            "attack_targets": attack_targets,
        }
        # 亡语召唤（V3-3，可选）：仅当填了 death_spawn_unit 才写入。
        ds_unit = _text(row.get("death_spawn_unit"))
        if ds_unit:
            units[unit_id]["death_spawn_unit"] = ds_unit
            units[unit_id]["death_spawn_count"] = int(_number(row.get("death_spawn_count"), f"unit {unit_id}.death_spawn_count"))

    cards: dict[str, Any] = {}
    card_enabled: dict[str, bool] = {}
    for row in card_rows:
        card_id = _require_id(row, "card_id", "Cards")
        if card_id in cards:
            raise ConfigError(f"duplicate card_id: {card_id}")
        enabled = _bool(row.get("enabled"))
        card_enabled[card_id] = enabled
        if not enabled:
            continue
        cards[card_id] = {
            "name": _text(row.get("name")),
            "elixir_cost": _number(row.get("elixir_cost"), f"card {card_id}.elixir_cost"),
            "skills": [],
        }

    skills_by_card: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for row in skill_rows:
        card_id = _require_id(row, "card_id", "CardSkills")
        if card_id not in card_enabled:
            raise ConfigError(f"CardSkills row {row['_row']} references unknown card {card_id}")
        if not card_enabled[card_id]:
            continue
        order = int(_number(row.get("order"), f"skill {card_id}.order"))
        skill_type = _text(row.get("skill_type"))
        if skill_type not in SKILL_TYPES:
            raise ConfigError(f"skill {card_id} order {order} skill_type must be one of {SKILL_TYPES}")

        block: dict[str, Any] = {"type": skill_type}
        if skill_type == "spawn_unit":
            unit_id = _text(row.get("unit_id"))
            if unit_id not in units:
                raise ConfigError(f"skill {card_id} order {order} references unknown unit {unit_id}")
            block["unit_id"] = unit_id
            block["count"] = _number(row.get("count"), f"skill {card_id}.count")
        elif skill_type == "direct_damage":
            block["damage"] = _number(row.get("damage"), f"skill {card_id}.damage")
            target = _text(row.get("target"))
            if target not in TARGETS:
                raise ConfigError(f"skill {card_id} order {order} target must be one of {TARGETS}")
            block["target"] = target
        elif skill_type == "aoe_damage":
            radius = _number(row.get("radius"), f"skill {card_id}.radius")
            if float(radius) < 0.0:
                raise ConfigError(f"skill {card_id}.radius must be >= 0")
            block["radius"] = radius
            block["damage"] = _number(row.get("damage"), f"skill {card_id}.damage")
        elif skill_type == "aoe_heal":
            radius = _number(row.get("radius"), f"skill {card_id}.radius")
            if float(radius) < 0.0:
                raise ConfigError(f"skill {card_id}.radius must be >= 0")
            block["radius"] = radius
            block["damage"] = _number(row.get("damage"), f"skill {card_id}.damage")   # damage 字段复用为治疗量

        skills_by_card.setdefault(card_id, []).append((order, block))

    for card_id in cards:
        blocks = skills_by_card.get(card_id, [])
        if not blocks:
            raise ConfigError(f"card {card_id} has no skills")
        cards[card_id]["skills"] = [block for _, block in sorted(blocks, key=lambda item: item[0])]

    deck_by_level_side: dict[tuple[str, str], list[str]] = {}
    for row in deck_rows:
        level_id = _require_id(row, "level_id", "Decks")
        side = _text(row.get("side"))
        if side not in SIDES:
            raise ConfigError(f"Decks row {row['_row']} side must be one of {SIDES}")
        key = (level_id, side)
        if key in deck_by_level_side:
            raise ConfigError(f"duplicate deck: level={level_id}, side={side}")
        deck = [_text(row.get(f"slot_{i}")) for i in range(1, 9)]
        if any(not card_id for card_id in deck):
            raise ConfigError(f"deck {level_id}/{side} must fill slot_1..slot_8")
        for card_id in deck:
            if card_id not in cards:
                raise ConfigError(f"deck {level_id}/{side} references unknown or disabled card {card_id}")
        deck_by_level_side[key] = deck

    levels: dict[str, Any] = {}
    for row in level_rows:
        level_id = _require_id(row, "level_id", "Levels")
        if level_id in levels:
            raise ConfigError(f"duplicate level_id: {level_id}")
        difficulty = _text(row.get("ai_difficulty"))
        if difficulty not in DIFFICULTIES:
            raise ConfigError(f"level {level_id} ai_difficulty must be one of {DIFFICULTIES}")
        player_deck = deck_by_level_side.get((level_id, "player"))
        ai_deck = deck_by_level_side.get((level_id, "ai"))
        if player_deck is None or ai_deck is None:
            raise ConfigError(f"level {level_id} must have player and ai decks")
        levels[level_id] = {
            "name": _text(row.get("name")),
            "elixir_regen_rate": _number_float(row.get("elixir_regen_rate"), f"level {level_id}.elixir_regen_rate"),
            "elixir_max": _number(row.get("elixir_max"), f"level {level_id}.elixir_max"),
            "match_duration": _number(row.get("match_duration"), f"level {level_id}.match_duration"),
            "player_deck": player_deck,
            "ai_deck": ai_deck,
            "ai_difficulty": difficulty,
            "tower_hp": {
                "king": _number(row.get("king_tower_hp"), f"level {level_id}.king_tower_hp"),
                "princess": _number(row.get("princess_tower_hp"), f"level {level_id}.princess_tower_hp"),
            },
        }

    return cards, units, levels


def write_json_outputs(cards: dict[str, Any], units: dict[str, Any], levels: dict[str, Any]) -> None:
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    _write_json(CARDS_JSON, cards)
    _write_json(UNITS_JSON, units)
    _write_json(LEVELS_JSON, levels)


def _setup_sheet(ws, headers: list[str], widths: dict[str, int] | None = None) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = (widths or {}).get(header, max(12, len(header) + 2))


def _add_list_validation(ws, column_name: str, values: list[str], max_row: int = 500) -> None:
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        return
    col = headers.index(column_name) + 1
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{max_row}")


def workbook_from_json(workbook_path: Path = WORKBOOK_PATH) -> None:
    cards = _load_json(CARDS_JSON)
    units = _load_json(UNITS_JSON)
    levels = _load_json(LEVELS_JSON)

    wb = Workbook()
    wb.remove(wb.active)

    ws_units = wb.create_sheet("Units")
    _setup_sheet(ws_units, UNIT_HEADERS, {"unit_id": 18, "name": 14, "notes": 26})
    for unit_id, unit in units.items():
        ws_units.append(
            [
                unit_id,
                unit.get("name", ""),
                unit.get("hp", ""),
                unit.get("damage", ""),
                unit.get("attack_speed", ""),
                unit.get("move_speed", ""),
                unit.get("attack_range", ""),
                unit.get("aggro_radius", ""),
                unit.get("body_radius", ""),
                unit.get("target_type", ""),
                unit.get("attack_targets", ""),
                unit.get("death_spawn_unit", ""),
                unit.get("death_spawn_count", ""),
                "",
            ]
        )
    _add_list_validation(ws_units, "unit_type", UNIT_TYPES)
    _add_list_validation(ws_units, "attack_targets", ATTACK_TARGETS)

    ws_cards = wb.create_sheet("Cards")
    _setup_sheet(ws_cards, CARD_HEADERS, {"card_id": 16, "name": 14, "category": 14, "notes": 26})
    for card_id, card in cards.items():
        category = "troop" if any(skill.get("type") == "spawn_unit" for skill in card.get("skills", [])) else "spell"
        ws_cards.append([card_id, card.get("name", ""), card.get("elixir_cost", ""), category, True, ""])

    ws_skills = wb.create_sheet("CardSkills")
    _setup_sheet(ws_skills, SKILL_HEADERS, {"card_id": 16, "skill_type": 16, "unit_id": 18, "target": 22, "notes": 26})
    for card_id, card in cards.items():
        for index, skill in enumerate(card.get("skills", []), start=1):
            ws_skills.append(
                [
                    card_id,
                    index,
                    skill.get("type", ""),
                    skill.get("unit_id", ""),
                    skill.get("count", ""),
                    skill.get("damage", ""),
                    skill.get("radius", ""),
                    skill.get("target", ""),
                    "",
                ]
            )
    _add_list_validation(ws_skills, "skill_type", SKILL_TYPES)
    _add_list_validation(ws_skills, "target", TARGETS)

    ws_levels = wb.create_sheet("Levels")
    _setup_sheet(ws_levels, LEVEL_HEADERS, {"level_id": 16, "name": 14, "ai_difficulty": 16, "notes": 26})
    for level_id, level in levels.items():
        tower_hp = level.get("tower_hp", {})
        ws_levels.append(
            [
                level_id,
                level.get("name", ""),
                level.get("elixir_regen_rate", ""),
                level.get("elixir_max", ""),
                level.get("match_duration", ""),
                level.get("ai_difficulty", ""),
                tower_hp.get("king", ""),
                tower_hp.get("princess", ""),
                "",
            ]
        )
    _add_list_validation(ws_levels, "ai_difficulty", DIFFICULTIES)

    ws_decks = wb.create_sheet("Decks")
    _setup_sheet(ws_decks, DECK_HEADERS, {"level_id": 16, "side": 12, "notes": 26})
    for level_id, level in levels.items():
        ws_decks.append([level_id, "player", *level.get("player_deck", []), ""])
        ws_decks.append([level_id, "ai", *level.get("ai_deck", []), ""])
    _add_list_validation(ws_decks, "side", SIDES)

    ws_balance = wb.create_sheet("Balance_View")
    _setup_sheet(
        ws_balance,
        [
            "unit_id",
            "name",
            "hp",
            "damage",
            "attack_interval_s",
            "dps",
            "move_speed",
            "attack_range",
            "unit_type",
        ],
        {"unit_id": 18, "name": 14},
    )
    for row_idx, (unit_id, unit) in enumerate(units.items(), start=2):
        ws_balance.append(
            [
                unit_id,
                unit.get("name", ""),
                unit.get("hp", ""),
                unit.get("damage", ""),
                unit.get("attack_speed", ""),
                f'=IF(E{row_idx}>0,D{row_idx}/E{row_idx},"")',
                unit.get("move_speed", ""),
                unit.get("attack_range", ""),
                unit.get("target_type", ""),
            ]
        )

    ws_enums = wb.create_sheet("_Enums")
    _setup_sheet(ws_enums, ["skill_type", "unit_type", "target", "side", "ai_difficulty"])
    max_len = max(len(SKILL_TYPES), len(UNIT_TYPES), len(TARGETS), len(SIDES), len(DIFFICULTIES))
    for i in range(max_len):
        ws_enums.append(
            [
                SKILL_TYPES[i] if i < len(SKILL_TYPES) else "",
                UNIT_TYPES[i] if i < len(UNIT_TYPES) else "",
                TARGETS[i] if i < len(TARGETS) else "",
                SIDES[i] if i < len(SIDES) else "",
                DIFFICULTIES[i] if i < len(DIFFICULTIES) else "",
            ]
        )
    ws_enums.sheet_state = "hidden"

    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def check_outputs(cards: dict[str, Any], units: dict[str, Any], levels: dict[str, Any]) -> None:
    current = {
        "cards.json": _load_json(CARDS_JSON),
        "units.json": _load_json(UNITS_JSON),
        "levels.json": _load_json(LEVELS_JSON),
    }
    generated = {"cards.json": cards, "units.json": units, "levels.json": levels}
    mismatches = [name for name in generated if generated[name] != current[name]]
    if mismatches:
        raise ConfigError("generated config differs from disk: " + ", ".join(mismatches))


def main() -> int:
    parser = argparse.ArgumentParser(description="Build or bootstrap game config.")
    parser.add_argument("--from-json", action="store_true", help="recreate config/GameConfig.xlsx from JSON files")
    parser.add_argument("--check", action="store_true", help="validate workbook and verify generated JSON matches disk")
    args = parser.parse_args()

    try:
        if args.from_json:
            workbook_from_json()
            print(f"wrote {WORKBOOK_PATH}")
            return 0

        cards, units, levels = build_json_from_workbook()
        if args.check:
            check_outputs(cards, units, levels)
            print("config check ok")
            return 0
        write_json_outputs(cards, units, levels)
        print(f"wrote {CARDS_JSON}")
        print(f"wrote {UNITS_JSON}")
        print(f"wrote {LEVELS_JSON}")
        return 0
    except ConfigError as exc:
        print(f"config error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
