#!/usr/bin/env python3
"""Build audio resource JSON from the designer-facing AudioConfig workbook.

Workflow:
    uv run --with openpyxl python tools/build_audio_config.py --init
    uv run --with openpyxl python tools/build_audio_config.py
    uv run --with openpyxl python tools/build_audio_config.py --check
    uv run --with openpyxl python tools/build_audio_config.py --from-json

Godot reads config/audio_assets.json at runtime. The workbook is the
human-readable audio plan and source table.
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl.comments import Comment
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / "config"
WORKBOOK_PATH = CONFIG_DIR / "AudioConfig.xlsx"
AUDIO_JSON = CONFIG_DIR / "audio_assets.json"

SHEET = "AudioAssets"
GUIDE_SHEET = "ColumnGuide"
ENUM_SHEET = "_Enums"

HEADERS = [
    "asset_id",
    "display_name_zh",
    "type",
    "group",
    "bus",
    "path",
    "asset_status",
    "loop",
    "volume_db",
    "pitch_min",
    "pitch_max",
    "max_polyphony",
    "priority",
    "suggested_duration_s",
    "implementation_phase",
    "trigger",
    "effect_notes",
    "source_notes",
]

TYPES = ["music", "ambience", "stinger", "ui", "sfx"]
BUSES = ["Music", "Ambience", "SFX", "UI"]
PHASES = ["P0", "P1", "P2"]
STATUSES = ["planned", "sourced", "imported", "final"]

COLUMN_INFO = {
    "asset_id": ("资源ID", "稳定运行时 ID，代码和 JSON 用它查找音频；不要随意改名。"),
    "display_name_zh": ("中文资源名", "给策划和音频制作看的中文名称，不参与代码查找。"),
    "type": ("资源类型", "music=音乐，ambience=环境底噪，stinger=短乐句，ui=界面音，sfx=游戏音效。"),
    "group": ("功能分组", "按使用场景粗分，如 battle、spell、tower、reward，方便筛选制作。"),
    "bus": ("音频总线", "未来接 Godot Audio Bus/音量设置用；当前不存在对应 bus 时会回退到 Master。"),
    "path": ("目标路径", "Godot res:// 路径，不代表文件已经存在；res://sound/... 对应工程根目录 sound/...。"),
    "asset_status": ("素材状态", "planned=仅列需求，sourced=已找到候选，imported=已入工程，final=已验收定稿。"),
    "loop": ("是否循环", "音乐和环境声通常为 TRUE，一次性音效为 FALSE。"),
    "volume_db": ("默认音量dB", "该资源默认播放音量；负数更小，后续可被总线/设置覆盖。"),
    "pitch_min": ("最低随机音高", "SFX 播放时的随机音高下限，用于减少重复感；音乐通常为 1.0。"),
    "pitch_max": ("最高随机音高", "SFX 播放时的随机音高上限；必须大于等于 pitch_min。"),
    "max_polyphony": ("最大并发", "同一资源最多同时播放多少个实例，用来限制高频命中音。"),
    "priority": ("优先级", "制作和混音优先级，数字越高越关键。"),
    "suggested_duration_s": ("建议时长秒", "给制作/选材的目标时长范围，不是运行时裁切规则。"),
    "implementation_phase": ("实现阶段", "P0=首批必须，P1=体验增强，P2=后续可选。"),
    "trigger": ("触发时机", "说明哪个场景/事件会播放它，便于查代码和验收。"),
    "effect_notes": ("效果说明", "中文声音设计说明：情绪、材质、层次、不能遮挡什么反馈。"),
    "source_notes": ("来源备注", "记录是否未入库、待采购、来源许可、候选包名等。"),
}


class AudioConfigError(RuntimeError):
    pass


def _row(
    asset_id: str,
    type_: str,
    group: str,
    bus: str,
    path: str,
    loop: bool,
    volume_db: float,
    pitch_min: float,
    pitch_max: float,
    max_polyphony: int,
    priority: int,
    suggested_duration_s: str,
    phase: str,
    trigger: str,
    effect_notes: str,
    source_notes: str = "未入库；待制作/采购音频文件",
) -> dict[str, Any]:
    return {
        "asset_id": asset_id,
        "display_name_zh": "",
        "type": type_,
        "group": group,
        "bus": bus,
        "path": path,
        "asset_status": "planned",
        "loop": loop,
        "volume_db": volume_db,
        "pitch_min": pitch_min,
        "pitch_max": pitch_max,
        "max_polyphony": max_polyphony,
        "priority": priority,
        "suggested_duration_s": suggested_duration_s,
        "implementation_phase": phase,
        "trigger": trigger,
        "effect_notes": effect_notes,
        "source_notes": source_notes,
    }


DEFAULT_ROWS: list[dict[str, Any]] = [
    _row("music_main_menu", "music", "menu", "Music", "res://sound/bgm/music_main_menu.ogg", True, -9.0, 1.0, 1.0, 1, 80, "45-75", "P0", "main_menu _ready", "dark medieval identity theme; low strings, frame drum, distant bell, subtle choir"),
    _row("music_battle_normal", "music", "battle", "Music", "res://sound/bgm/music_battle_normal.ogg", True, -10.0, 1.0, 1.0, 1, 90, "60-90", "P0", "normal battle _ready", "mid-tempo combat loop; drums clear but sparse, leaves room for SFX"),
    _row("music_battle_boss", "music", "battle", "Music", "res://sound/bgm/music_battle_boss.ogg", True, -9.0, 1.0, 1.0, 1, 95, "60-90", "P0", "elite/boss battle _ready", "more pressure than normal battle; low drums, dissonant strings, dark choir hits"),
    _row("music_run_map", "music", "run", "Music", "res://sound/bgm/music_run_map.ogg", True, -11.0, 1.0, 1.0, 1, 70, "45-75", "P1", "run_scene _ready", "journey/map loop; restrained, ominous, campfire and road mood"),
    _row("music_reward_ambient", "music", "reward", "Music", "res://sound/bgm/music_reward_ambient.ogg", True, -13.0, 1.0, 1.0, 1, 45, "20-40", "P1", "reward overlay open", "mystic reward bed; optional if run map music already carries reward screens"),
    _row("music_tutorial", "music", "campaign", "Music", "res://sound/bgm/music_tutorial.ogg", True, -12.0, 1.0, 1.0, 1, 40, "45-75", "P2", "future tutorial/campaign", "quieter teaching loop; less tension than combat"),
    _row("amb_battle_wind", "ambience", "battle", "Ambience", "res://sound/ambience/amb_battle_wind.ogg", True, -22.0, 1.0, 1.0, 1, 25, "30-60", "P1", "battle ambience layer", "low wind, distant banners, very quiet; must not mask gameplay feedback"),
    _row("amb_run_campfire", "ambience", "run", "Ambience", "res://sound/ambience/amb_run_campfire.ogg", True, -24.0, 1.0, 1.0, 1, 20, "30-60", "P2", "run map ambience layer", "campfire, distant bell, night air; optional under run map music"),

    _row("stinger_battle_start", "stinger", "battle", "SFX", "res://sound/stingers/stinger_battle_start.wav", False, -5.0, 1.0, 1.0, 1, 75, "1-2", "P1", "battle starts", "short drum pickup / horn breath before first action"),
    _row("stinger_victory", "stinger", "result", "SFX", "res://sound/stingers/stinger_victory.wav", False, -4.0, 1.0, 1.0, 1, 100, "2-4", "P0", "player win result", "clear victory cadence; bronze bell, small choir lift, no long tail"),
    _row("stinger_defeat", "stinger", "result", "SFX", "res://sound/stingers/stinger_defeat.wav", False, -4.0, 1.0, 1.0, 1, 100, "2-4", "P0", "player lose result", "descending low bell and drum; dark but not annoying"),
    _row("stinger_draw", "stinger", "result", "SFX", "res://sound/stingers/stinger_draw.wav", False, -5.0, 1.0, 1.0, 1, 85, "1.5-3", "P1", "draw result", "neutral unresolved cadence"),
    _row("stinger_run_cleared", "stinger", "result", "SFX", "res://sound/stingers/stinger_run_cleared.wav", False, -3.0, 1.0, 1.0, 1, 100, "4-8", "P1", "roguelite run won", "bigger win sting for full run clear; still short enough for repeat play"),

    _row("ui_button_press", "ui", "ui", "UI", "res://sound/ui/ui_button_press.wav", False, -9.0, 0.98, 1.03, 4, 45, "0.05-0.15", "P0", "all button pressed", "dry wood/stone click; fast UI confirmation"),
    _row("ui_button_back", "ui", "ui", "UI", "res://sound/ui/ui_button_back.wav", False, -10.0, 0.98, 1.02, 2, 40, "0.08-0.20", "P1", "back/menu buttons", "lower click/swipe back"),
    _row("ui_panel_open", "ui", "ui", "UI", "res://sound/ui/ui_panel_open.wav", False, -10.0, 1.0, 1.0, 2, 40, "0.15-0.35", "P1", "settings/reward/result panel appears", "cloth/parchment whoosh with small chime"),
    _row("ui_card_pickup", "ui", "card", "UI", "res://sound/ui/ui_card_pickup.wav", False, -8.0, 0.98, 1.04, 4, 65, "0.05-0.12", "P0", "card drag begins", "card lifts from hand; crisp parchment plus tiny magic tick"),
    _row("ui_card_drop_valid", "ui", "card", "UI", "res://sound/ui/ui_card_drop_valid.wav", False, -8.0, 0.98, 1.04, 4, 70, "0.08-0.20", "P0", "card successfully played", "positive snap into field; pairs with deploy/spell SFX"),
    _row("ui_card_drop_invalid", "ui", "card", "UI", "res://sound/ui/ui_card_drop_invalid.wav", False, -9.0, 0.98, 1.02, 3, 70, "0.08-0.20", "P0", "illegal drop or no elixir", "short dull blocked sound; avoid sharp error beep"),
    _row("ui_card_cancel", "ui", "card", "UI", "res://sound/ui/ui_card_cancel.wav", False, -10.0, 0.98, 1.02, 3, 35, "0.05-0.15", "P1", "card released outside field", "soft card return"),
    _row("ui_settings_toggle", "ui", "settings", "UI", "res://sound/ui/ui_settings_toggle.wav", False, -11.0, 1.0, 1.0, 2, 25, "0.05-0.15", "P2", "settings toggles", "small switch/click"),
    _row("ui_language_switch", "ui", "settings", "UI", "res://sound/ui/ui_language_switch.wav", False, -10.0, 1.0, 1.0, 1, 25, "0.15-0.30", "P2", "language changed", "tiny magical page turn"),
    _row("ui_volume_slide", "ui", "settings", "UI", "res://sound/ui/ui_volume_slide.wav", False, -14.0, 1.0, 1.0, 2, 20, "0.03-0.08", "P2", "future volume slider", "very subtle tick; rate-limit in code if needed"),

    _row("elixir_spend", "sfx", "elixir", "SFX", "res://sound/sfx/elixir_spend.wav", False, -8.0, 0.98, 1.04, 4, 70, "0.10-0.25", "P0", "card cost paid", "magic liquid drain; confirms resource spend"),
    _row("elixir_insufficient", "sfx", "elixir", "SFX", "res://sound/sfx/elixir_insufficient.wav", False, -8.0, 1.0, 1.0, 2, 80, "0.10-0.20", "P0", "cannot afford card", "dry empty vessel knock; distinct from invalid placement"),
    _row("elixir_full", "sfx", "elixir", "SFX", "res://sound/sfx/elixir_full.wav", False, -11.0, 1.0, 1.0, 1, 45, "0.25-0.50", "P1", "elixir reaches max", "subtle magic bloom; not every frame, cooldown required"),
    _row("elixir_tick", "sfx", "elixir", "SFX", "res://sound/sfx/elixir_tick.wav", False, -20.0, 0.98, 1.02, 1, 10, "0.03-0.08", "P2", "optional individual pip gain", "very quiet; may be disabled if annoying"),

    _row("run_node_select", "ui", "run", "UI", "res://sound/ui/run_node_select.wav", False, -9.0, 1.0, 1.0, 2, 35, "0.10-0.25", "P1", "run node highlighted/selected", "map token click on parchment"),
    _row("run_node_complete", "ui", "run", "UI", "res://sound/ui/run_node_complete.wav", False, -8.0, 1.0, 1.0, 1, 55, "0.25-0.60", "P1", "node marked completed", "wax seal stamp / small victory mark"),
    _row("run_boss_reveal", "ui", "run", "UI", "res://sound/ui/run_boss_reveal.wav", False, -6.0, 1.0, 1.0, 1, 70, "0.50-1.20", "P1", "boss node/reward appears", "ominous boom plus bell"),
    _row("reward_panel_open", "ui", "reward", "UI", "res://sound/ui/reward_panel_open.wav", False, -8.0, 1.0, 1.0, 1, 55, "0.25-0.60", "P0", "reward panel opens", "parchment unfurl + magical shimmer"),
    _row("reward_card_reveal", "ui", "reward", "UI", "res://sound/ui/reward_card_reveal.wav", False, -9.0, 0.97, 1.03, 3, 55, "0.15-0.35", "P1", "each draft card pops in", "short card flip; stagger-friendly"),
    _row("reward_card_pick", "ui", "reward", "UI", "res://sound/ui/reward_card_pick.wav", False, -7.0, 1.0, 1.0, 1, 75, "0.35-0.80", "P0", "draft card chosen", "positive magic lock-in; should feel valuable"),
    _row("reward_skip", "ui", "reward", "UI", "res://sound/ui/reward_skip.wav", False, -10.0, 1.0, 1.0, 1, 30, "0.10-0.25", "P1", "reward skipped", "soft close, no negative sting"),
    _row("relic_reveal", "ui", "reward", "UI", "res://sound/ui/relic_reveal.wav", False, -7.0, 1.0, 1.0, 2, 75, "0.45-0.90", "P0", "relic offer appears", "ancient object shimmer; more ceremonial than cards"),
    _row("relic_pick", "ui", "reward", "UI", "res://sound/ui/relic_pick.wav", False, -6.0, 1.0, 1.0, 1, 85, "0.50-1.00", "P0", "relic chosen", "dark magical seal; player should remember it"),
    _row("meta_unlock", "ui", "run", "UI", "res://sound/ui/meta_unlock.wav", False, -6.0, 1.0, 1.0, 1, 85, "0.80-1.50", "P1", "new relic/meta unlock", "larger unlock flourish; rare and satisfying"),

    _row("deploy_small", "sfx", "deploy", "SFX", "res://sound/sfx/deploy_small.wav", False, -8.0, 0.94, 1.08, 6, 65, "0.12-0.25", "P0", "small troop groups spawn", "light foot/cloth pop for goblins, skeletons, minions"),
    _row("deploy_medium", "sfx", "deploy", "SFX", "res://sound/sfx/deploy_medium.wav", False, -7.0, 0.96, 1.05, 4, 65, "0.15-0.35", "P0", "standard single troop spawn", "armor/boot landing for knight, archer, mage"),
    _row("deploy_large", "sfx", "deploy", "SFX", "res://sound/sfx/deploy_large.wav", False, -6.0, 0.94, 1.03, 3, 75, "0.25-0.55", "P0", "giant/golem spawn", "heavy ground thump and low armor/stone resonance"),
    _row("deploy_air", "sfx", "deploy", "SFX", "res://sound/sfx/deploy_air.wav", False, -8.0, 0.96, 1.05, 4, 60, "0.20-0.45", "P1", "flying troop spawn", "spectral lift / wingless whoosh for wraiths and fire skull"),
    _row("deploy_spell_cast", "sfx", "deploy", "SFX", "res://sound/sfx/deploy_spell_cast.wav", False, -8.0, 0.97, 1.03, 4, 60, "0.10-0.25", "P1", "generic spell release if card has no custom cast", "short magic cast bed"),
    _row("deploy_death_spawn", "sfx", "deploy", "SFX", "res://sound/sfx/deploy_death_spawn.wav", False, -7.0, 0.95, 1.04, 4, 75, "0.25-0.55", "P0", "death_spawn units appear", "corpse split / bone and shadow burst"),

    _row("attack_sword", "sfx", "combat", "SFX", "res://sound/sfx/attack_sword.wav", False, -9.0, 0.94, 1.07, 8, 45, "0.08-0.18", "P0", "knight/standard melee swing", "short metal slash; pair with hit impact"),
    _row("attack_blade_small", "sfx", "combat", "SFX", "res://sound/sfx/attack_blade_small.wav", False, -10.0, 0.94, 1.10, 10, 40, "0.05-0.14", "P1", "goblin/skeleton fast melee", "light stab/clack variation"),
    _row("attack_heavy", "sfx", "combat", "SFX", "res://sound/sfx/attack_heavy.wav", False, -7.0, 0.92, 1.04, 5, 60, "0.12-0.28", "P0", "berserker/ogre/golem heavy swing", "large whoosh and weight before impact"),
    _row("hit_light", "sfx", "combat", "SFX", "res://sound/sfx/hit_light.wav", False, -11.0, 0.95, 1.10, 12, 35, "0.04-0.12", "P0", "minor unit damage", "soft flesh/cloth tick; frequently repeated"),
    _row("hit_medium", "sfx", "combat", "SFX", "res://sound/sfx/hit_medium.wav", False, -9.0, 0.94, 1.08, 10, 45, "0.08-0.18", "P0", "normal unit damage", "meaty hit with small armor element"),
    _row("hit_heavy", "sfx", "combat", "SFX", "res://sound/sfx/hit_heavy.wav", False, -7.0, 0.93, 1.05, 6, 65, "0.15-0.35", "P0", "big hit / fireball / heavy unit", "low impact; supports hitstop/screenshake"),
    _row("hit_armor", "sfx", "combat", "SFX", "res://sound/sfx/hit_armor.wav", False, -9.0, 0.94, 1.06, 8, 45, "0.08-0.20", "P1", "armored unit hit", "metal clang without being too bright"),
    _row("hit_tower", "sfx", "tower", "SFX", "res://sound/sfx/hit_tower.wav", False, -8.0, 0.96, 1.04, 6, 55, "0.10-0.25", "P0", "tower takes damage", "stone/wood crack; distinct from unit hits"),
    _row("death_small", "sfx", "death", "SFX", "res://sound/sfx/death_small.wav", False, -10.0, 0.95, 1.08, 8, 45, "0.15-0.35", "P0", "small living unit dies", "quick fall/cloth, no gore emphasis"),
    _row("death_bone", "sfx", "death", "SFX", "res://sound/sfx/death_bone.wav", False, -9.0, 0.95, 1.08, 8, 45, "0.15-0.35", "P0", "skeleton dies", "bone clatter"),
    _row("death_spirit", "sfx", "death", "SFX", "res://sound/sfx/death_spirit.wav", False, -10.0, 0.96, 1.04, 6, 45, "0.20-0.45", "P1", "wraith/fire skull dies", "spectral dissolve / smoky vanish"),
    _row("death_large", "sfx", "death", "SFX", "res://sound/sfx/death_large.wav", False, -6.0, 0.94, 1.03, 4, 70, "0.35-0.80", "P0", "large unit dies", "heavy body/stone collapse"),

    _row("bow_shot", "sfx", "projectile", "SFX", "res://sound/sfx/bow_shot.wav", False, -10.0, 0.96, 1.08, 8, 45, "0.05-0.15", "P0", "archer projectile fired", "bow twang and arrow hiss"),
    _row("arrow_hit", "sfx", "projectile", "SFX", "res://sound/sfx/arrow_hit.wav", False, -11.0, 0.95, 1.08, 10, 35, "0.05-0.15", "P1", "arrow projectile impact", "small stick/thud"),
    _row("magic_bolt_cast", "sfx", "projectile", "SFX", "res://sound/sfx/magic_bolt_cast.wav", False, -9.0, 0.96, 1.06, 6, 50, "0.08-0.20", "P0", "musketeer/sorceress bolt fired", "dark purple magical snap"),
    _row("magic_bolt_hit", "sfx", "projectile", "SFX", "res://sound/sfx/magic_bolt_hit.wav", False, -9.0, 0.96, 1.06, 6, 45, "0.08-0.20", "P1", "magic bolt impact", "small arcane pop"),
    _row("fire_skull_shot", "sfx", "projectile", "SFX", "res://sound/sfx/fire_skull_shot.wav", False, -8.0, 0.96, 1.04, 5, 55, "0.10-0.25", "P1", "baby_dragon/fire skull projectile", "small flame spit"),
    _row("tower_princess_shot", "sfx", "tower", "SFX", "res://sound/sfx/tower_princess_shot.wav", False, -9.0, 0.96, 1.06, 4, 60, "0.08-0.18", "P0", "princess tower fires", "crossbow/ballista snap"),
    _row("tower_king_shot", "sfx", "tower", "SFX", "res://sound/sfx/tower_king_shot.wav", False, -8.0, 0.96, 1.04, 3, 70, "0.10-0.25", "P1", "king tower fires", "heavier tower shot"),
    _row("tower_destroy_princess", "sfx", "tower", "SFX", "res://sound/sfx/tower_destroy_princess.wav", False, -4.0, 1.0, 1.0, 1, 90, "0.8-1.5", "P0", "princess tower destroyed", "stone collapse, dust, short fire burst"),
    _row("tower_destroy_king", "sfx", "tower", "SFX", "res://sound/sfx/tower_destroy_king.wav", False, -3.0, 1.0, 1.0, 1, 100, "1.2-2.5", "P0", "king tower destroyed", "larger castle collapse; transitions into result sting"),

    _row("spell_fireball_cast", "sfx", "spell", "SFX", "res://sound/sfx/spell_fireball_cast.wav", False, -7.0, 0.98, 1.03, 4, 75, "0.15-0.35", "P0", "fireball card released", "flame ignition and throw"),
    _row("spell_fireball_fly", "sfx", "spell", "SFX", "res://sound/sfx/spell_fireball_fly.wav", False, -11.0, 0.98, 1.02, 3, 35, "0.25-0.60", "P1", "fireball projectile travel", "short flame whoosh; can be skipped if travel is instant"),
    _row("spell_fireball_impact", "sfx", "spell", "SFX", "res://sound/sfx/spell_fireball_impact.wav", False, -4.0, 0.96, 1.03, 3, 95, "0.5-1.0", "P0", "fireball explosion", "deep explosion, crackle tail; high gameplay readability"),
    _row("spell_arrows_cast", "sfx", "spell", "SFX", "res://sound/sfx/spell_arrows_cast.wav", False, -8.0, 0.96, 1.04, 4, 65, "0.12-0.30", "P0", "arrows card released", "volley release / bow cluster"),
    _row("spell_arrows_impact", "sfx", "spell", "SFX", "res://sound/sfx/spell_arrows_impact.wav", False, -8.0, 0.96, 1.05, 4, 70, "0.35-0.70", "P0", "arrows hit area", "many arrows striking dirt/armor"),
    _row("spell_zap_cast", "sfx", "spell", "SFX", "res://sound/sfx/spell_zap_cast.wav", False, -7.0, 0.98, 1.04, 4, 80, "0.08-0.20", "P0", "zap card released", "tight electric snap"),
    _row("spell_zap_impact", "sfx", "spell", "SFX", "res://sound/sfx/spell_zap_impact.wav", False, -6.0, 0.98, 1.05, 4, 85, "0.12-0.30", "P0", "zap direct hit", "short crackle impact"),
    _row("spell_lightning_cast", "sfx", "spell", "SFX", "res://sound/sfx/spell_lightning_cast.wav", False, -6.0, 0.98, 1.02, 3, 80, "0.20-0.45", "P0", "lightning card released", "air charge before strike"),
    _row("spell_lightning_impact", "sfx", "spell", "SFX", "res://sound/sfx/spell_lightning_impact.wav", False, -3.5, 0.98, 1.02, 3, 95, "0.45-0.90", "P0", "lightning hit", "large thunder crack with quick tail"),
    _row("spell_log_roll", "sfx", "spell", "SFX", "res://sound/sfx/spell_log_roll.wav", False, -9.0, 0.96, 1.04, 3, 55, "0.35-0.80", "P1", "rolling boulder/log travel", "stone rolling over dirt/cobble"),
    _row("spell_log_impact", "sfx", "spell", "SFX", "res://sound/sfx/spell_log_impact.wav", False, -6.0, 0.96, 1.04, 4, 80, "0.25-0.55", "P0", "rolling boulder impact", "stone crunch and body knock"),
    _row("spell_heal_cast", "sfx", "spell", "SFX", "res://sound/sfx/spell_heal_cast.wav", False, -8.0, 0.98, 1.02, 3, 70, "0.20-0.45", "P0", "heal card released", "soft sacred/green magic open"),
    _row("spell_heal_tick", "sfx", "spell", "SFX", "res://sound/sfx/spell_heal_tick.wav", False, -12.0, 0.98, 1.04, 6, 35, "0.10-0.25", "P1", "heal amount applied", "gentle shimmer; use sparingly if repeated"),
    _row("spell_heal_end", "sfx", "spell", "SFX", "res://sound/sfx/spell_heal_end.wav", False, -10.0, 1.0, 1.0, 2, 35, "0.20-0.40", "P2", "future heal over-time ends", "soft close shimmer"),
    _row("spell_golem_death_spawn", "sfx", "spell", "SFX", "res://sound/sfx/spell_golem_death_spawn.wav", False, -5.0, 0.96, 1.03, 2, 90, "0.6-1.2", "P0", "golem death_spawn triggers", "large collapse plus smaller creatures breaking out"),
]

ZH_AUDIO_META: dict[str, tuple[str, str]] = {
    "music_main_menu": ("主菜单主题曲", "黑暗中世纪身份主题；低弦乐、框鼓、远钟和微弱唱诗，先把世界观立住。"),
    "music_battle_normal": ("常规战斗循环", "中速战斗循环；鼓点清楚但不要过密，给部署、命中和法术音效留空间。"),
    "music_battle_boss": ("精英/首领战斗循环", "比常规战斗更压迫；低鼓、紧张弦乐、不协和合唱点缀，突出高风险节点。"),
    "music_run_map": ("Roguelite 地图音乐", "旅途/节点地图循环；克制、阴郁，有篝火、夜路和远方钟声的感觉。"),
    "music_reward_ambient": ("奖励界面氛围音乐", "神秘奖励底音；如果 run 地图音乐已经能承托奖励界面，可后续取消或弱化。"),
    "music_tutorial": ("教学/短战役音乐", "教学用安静循环；紧张度低于战斗音乐，避免干扰新手理解操作。"),
    "amb_battle_wind": ("战场风声环境", "低音量风声、远处旗帜和空旷战场气氛；必须很轻，不能盖住玩法反馈。"),
    "amb_run_campfire": ("营地篝火环境", "篝火、远钟、夜风；作为 run 地图音乐下的可选环境层。"),
    "stinger_battle_start": ("战斗开始短乐句", "短鼓点或号角吸气，提示对局开始；不要拖太长，避免压住首轮出牌。"),
    "stinger_victory": ("胜利短乐句", "清晰胜利收束；青铜钟、小合唱上扬，尾音短，适合频繁对局。"),
    "stinger_defeat": ("失败短乐句", "下行低钟和鼓点；黑暗但不要刺耳，避免失败后反复听产生烦躁。"),
    "stinger_draw": ("平局短乐句", "中性、未解决的收束；表达没赢也没输的悬置感。"),
    "stinger_run_cleared": ("通关胜利短乐句", "整局通关的大胜利乐句；比普通胜利更隆重，但仍要短，适合重复游玩。"),
    "ui_button_press": ("按钮点击", "干燥木头/石质点击，快速确认 UI 操作。"),
    "ui_button_back": ("返回按钮", "更低、更软的返回点击或轻扫声，和普通确认点击区分。"),
    "ui_panel_open": ("面板打开", "布料/羊皮纸展开加轻微提示音，用于设置、奖励、结算等面板。"),
    "ui_card_pickup": ("卡牌拿起", "卡牌从手牌抬起；清脆羊皮纸摩擦加很小的魔法点声。"),
    "ui_card_drop_valid": ("卡牌合法落下", "正向的卡牌落位声；和部署/法术主音效叠加后仍要干净。"),
    "ui_card_drop_invalid": ("卡牌非法落下", "短促沉闷的受阻声；避免做成刺耳报错音。"),
    "ui_card_cancel": ("卡牌取消拖拽", "柔和的卡牌回手声，提示没有扣费也没有出牌。"),
    "ui_settings_toggle": ("设置开关", "很小的拨片/开关点击，用于设置项。"),
    "ui_language_switch": ("语言切换", "轻微魔法翻页声，提示语言配置已切换。"),
    "ui_volume_slide": ("音量滑条刻度", "极轻的刻度 tick；如果后续接滑条，需要在代码里限频。"),
    "elixir_spend": ("圣水消耗", "魔法液体被抽走的声音，确认资源已支付。"),
    "elixir_insufficient": ("圣水不足", "空容器的干敲声；要和非法落点音区分。"),
    "elixir_full": ("圣水已满", "轻微魔法充盈声；必须加冷却，不能满槽后反复触发。"),
    "elixir_tick": ("圣水单格增长", "很轻的单格增长提示；如果听感烦，可默认关闭。"),
    "run_node_select": ("地图节点选择", "地图棋子/标记落在羊皮纸上的点击声。"),
    "run_node_complete": ("地图节点完成", "蜡封盖章或小胜利标记声，确认节点已清。"),
    "run_boss_reveal": ("首领节点揭示", "阴沉重击加钟声，用于 boss 节点或 boss 奖励出现。"),
    "reward_panel_open": ("奖励面板打开", "羊皮纸展开加魔法微光，提示战后奖励出现。"),
    "reward_card_reveal": ("奖励卡牌揭示", "短卡牌翻面声，适合三张卡错峰出现。"),
    "reward_card_pick": ("选择奖励卡", "正向魔法锁定声；应让玩家觉得这次选择有价值。"),
    "reward_skip": ("跳过奖励", "柔和关闭声，不要带失败或惩罚情绪。"),
    "relic_reveal": ("遗物揭示", "古物微光声，比普通卡牌更有仪式感。"),
    "relic_pick": ("选择遗物", "黑暗魔法封印声；要让玩家记住这次强化。"),
    "meta_unlock": ("局间解锁", "较大的解锁音效；稀有、满足，但不要像胜利音乐一样拖长。"),
    "deploy_small": ("小型单位部署", "轻脚步、布料或骨头小弹出声，适合骷髅、哥布林、亡魂等小群体。"),
    "deploy_medium": ("中型单位部署", "盔甲和靴子落地声，适合骑士、弓手、女巫等标准单位。"),
    "deploy_large": ("大型单位部署", "重踏地面加低频盔甲/石头共鸣，突出巨人和魔像重量。"),
    "deploy_air": ("空中单位部署", "幽魂升起或气流声，适合飞兵和火颅，不要做成真实鸟翼。"),
    "deploy_spell_cast": ("通用法术释放", "短促魔法释放底声；仅给没有专属 cast 的法术兜底。"),
    "deploy_death_spawn": ("亡语召唤", "尸体裂开、骨片和暗影爆开的混合声，用于 death_spawn。"),
    "attack_sword": ("标准剑击挥砍", "短金属挥砍声；主要表达攻击动作，命中感交给 hit 音效。"),
    "attack_blade_small": ("小型快速刺击", "轻刺、短刀或骨爪声，适合哥布林/骷髅等快攻单位。"),
    "attack_heavy": ("重型挥击", "大幅度挥砍和重量感，命中前就能感觉到重击要来了。"),
    "hit_light": ("轻度受击", "轻微布料/身体命中声；会频繁播放，必须短且不抢耳。"),
    "hit_medium": ("中度受击", "肉感命中加少量盔甲元素，是大多数普通伤害的默认反馈。"),
    "hit_heavy": ("重度受击", "低频冲击，配合顿帧和震屏表达大伤害。"),
    "hit_armor": ("护甲受击", "金属碰撞但不要太亮，避免高频刺耳。"),
    "hit_tower": ("防御塔受击", "石头/木结构裂响，必须和单位受击区分。"),
    "death_small": ("小型单位死亡", "快速倒地和布料声，不强调血腥。"),
    "death_bone": ("骷髅死亡", "骨头散架和落地声。"),
    "death_spirit": ("幽魂死亡", "烟雾和灵体消散声，适合亡魂、火颅等。"),
    "death_large": ("大型单位死亡", "重物倒塌或石块坍塌，体现大型单位死亡重量。"),
    "bow_shot": ("弓箭发射", "弓弦声加箭矢掠风，用于弓手远程开火。"),
    "arrow_hit": ("箭矢命中", "小型插入或钝击声，作为箭矢落点反馈。"),
    "magic_bolt_cast": ("魔法弹发射", "暗紫色魔法弹的短促弹射声，用于女巫/远程法师。"),
    "magic_bolt_hit": ("魔法弹命中", "小型奥术爆点声。"),
    "fire_skull_shot": ("火颅投射物发射", "小火焰喷吐声，不能盖过火球法术。"),
    "tower_princess_shot": ("公主塔射击", "弩机或小型床弩弹射声。"),
    "tower_king_shot": ("王塔射击", "更厚重的塔射击声，和公主塔区分主次。"),
    "tower_destroy_princess": ("公主塔摧毁", "石块坍塌、尘土和短火光，配合塔毁震屏。"),
    "tower_destroy_king": ("王塔摧毁", "更大的城堡坍塌声，并自然衔接胜负短乐句。"),
    "spell_fireball_cast": ("火球释放", "火焰点燃并投出的声音。"),
    "spell_fireball_fly": ("火球飞行", "短火焰呼啸；如果当前表现层是瞬发，可后续不启用。"),
    "spell_fireball_impact": ("火球爆炸", "低频爆炸加火焰噼啪尾音，是高优先级玩法反馈。"),
    "spell_arrows_cast": ("箭雨释放", "一组弓弦齐发或箭矢出手声。"),
    "spell_arrows_impact": ("箭雨命中", "大量箭矢扎入地面、护甲和单位的密集命中声。"),
    "spell_zap_cast": ("电击释放", "紧凑电流起手声，动作要极短。"),
    "spell_zap_impact": ("电击命中", "短促电流爆点和噼啪声。"),
    "spell_lightning_cast": ("闪电释放", "雷击前的空气蓄能声。"),
    "spell_lightning_impact": ("闪电命中", "大雷击裂响加短尾音，必须清楚但不能盖住结算音。"),
    "spell_log_roll": ("滚石/滚木移动", "石头或木头滚过泥土、鹅卵石的声音。"),
    "spell_log_impact": ("滚石/滚木撞击", "石块挤压和单位被撞开的冲击声。"),
    "spell_heal_cast": ("治疗术释放", "柔和的圣洁/绿色魔法展开声。"),
    "spell_heal_tick": ("治疗生效", "温和微光声；如果治疗多次跳数，播放要克制。"),
    "spell_heal_end": ("治疗结束", "柔和收束的微光声，供后续持续治疗版本使用。"),
    "spell_golem_death_spawn": ("魔像亡语召唤", "大型坍塌后有小单位破出的声层，突出亡语触发。"),
}


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _num(value: Any, field: str) -> int | float:
    if value is None or value == "":
        raise AudioConfigError(f"missing number: {field}")
    if isinstance(value, bool):
        raise AudioConfigError(f"{field} must be a number, got boolean")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise AudioConfigError(f"{field} must be a number, got {value!r}") from exc
    if not math.isfinite(number):
        raise AudioConfigError(f"{field} must be finite")
    return int(number) if number.is_integer() else number


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or value == "":
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on", "是"}


def _read_rows(workbook_path: Path = WORKBOOK_PATH) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise AudioConfigError(f"missing sheet: {SHEET}")
    ws = wb[SHEET]
    actual = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
    if actual != HEADERS:
        raise AudioConfigError(f"{SHEET} headers mismatch: expected {HEADERS}, got {actual}")
    rows: list[dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        values = {HEADERS[i]: ws.cell(row=row_index, column=i + 1).value for i in range(len(HEADERS))}
        if all(_is_blank(v) for v in values.values()):
            continue
        values["_row"] = row_index
        rows.append(values)
    return rows


def rows_to_audio(rows: list[dict[str, Any]]) -> dict[str, Any]:
    audio: dict[str, Any] = {}
    for row in rows:
        row_ref = f"{SHEET}!row {row.get('_row', '?')}"
        asset_id = _text(row.get("asset_id"))
        if not asset_id:
            raise AudioConfigError(f"{row_ref} missing asset_id")
        if asset_id in audio:
            raise AudioConfigError(f"duplicate asset_id: {asset_id}")
        display_name_zh = _text(row.get("display_name_zh"))
        if not display_name_zh:
            raise AudioConfigError(f"{asset_id}.display_name_zh is required")
        type_ = _text(row.get("type"))
        if type_ not in TYPES:
            raise AudioConfigError(f"{asset_id}.type must be one of {TYPES}")
        bus = _text(row.get("bus"))
        if bus not in BUSES:
            raise AudioConfigError(f"{asset_id}.bus must be one of {BUSES}")
        asset_status = _text(row.get("asset_status"))
        if asset_status not in STATUSES:
            raise AudioConfigError(f"{asset_id}.asset_status must be one of {STATUSES}")
        phase = _text(row.get("implementation_phase"))
        if phase not in PHASES:
            raise AudioConfigError(f"{asset_id}.implementation_phase must be one of {PHASES}")
        path = _text(row.get("path"))
        if not path.startswith("res://sound/"):
            raise AudioConfigError(f"{asset_id}.path must be under res://sound/")
        pitch_min = float(_num(row.get("pitch_min"), f"{asset_id}.pitch_min"))
        pitch_max = float(_num(row.get("pitch_max"), f"{asset_id}.pitch_max"))
        if pitch_min <= 0.0 or pitch_max <= 0.0 or pitch_min > pitch_max:
            raise AudioConfigError(f"{asset_id}.pitch range must be positive and ordered")
        max_polyphony = int(_num(row.get("max_polyphony"), f"{asset_id}.max_polyphony"))
        if max_polyphony < 1:
            raise AudioConfigError(f"{asset_id}.max_polyphony must be >= 1")
        audio[asset_id] = {
            "display_name_zh": display_name_zh,
            "type": type_,
            "group": _text(row.get("group")),
            "bus": bus,
            "path": path,
            "asset_status": asset_status,
            "loop": _bool(row.get("loop")),
            "volume_db": float(_num(row.get("volume_db"), f"{asset_id}.volume_db")),
            "pitch_min": pitch_min,
            "pitch_max": pitch_max,
            "max_polyphony": max_polyphony,
            "priority": int(_num(row.get("priority"), f"{asset_id}.priority")),
            "suggested_duration_s": _text(row.get("suggested_duration_s")),
            "implementation_phase": phase,
            "trigger": _text(row.get("trigger")),
            "effect_notes": _text(row.get("effect_notes")),
            "source_notes": _text(row.get("source_notes")),
        }
    return audio


def _write_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _style_header(ws) -> None:
    fill = PatternFill(fill_type="solid", fgColor="263238")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell.value in COLUMN_INFO:
            zh_name, desc = COLUMN_INFO[str(cell.value)]
            cell.comment = Comment(f"{zh_name}\n{desc}", "Codex")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _set_widths(ws) -> None:
    widths = {
        "A": 28, "B": 20, "C": 12, "D": 14, "E": 12, "F": 42, "G": 14,
        "H": 9, "I": 11, "J": 10, "K": 10, "L": 14, "M": 10, "N": 18,
        "O": 16, "P": 32, "Q": 64, "R": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 42
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _add_column_guide(wb) -> None:
    ws = wb.create_sheet(GUIDE_SHEET)
    ws.append(["column", "中文名", "用途说明", "是否导出到 JSON"])
    for field in HEADERS:
        zh_name, desc = COLUMN_INFO[field]
        ws.append([field, zh_name, desc, "是"])
    _style_header(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 72
    ws.column_dimensions["D"].width = 16
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 34
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _add_validations(wb, ws) -> None:
    enum = wb.create_sheet(ENUM_SHEET)
    enum.sheet_state = "hidden"
    enum["A1"] = "type"
    enum["B1"] = "bus"
    enum["C1"] = "phase"
    enum["D1"] = "status"
    for i, value in enumerate(TYPES, start=2):
        enum.cell(i, 1).value = value
    for i, value in enumerate(BUSES, start=2):
        enum.cell(i, 2).value = value
    for i, value in enumerate(PHASES, start=2):
        enum.cell(i, 3).value = value
    for i, value in enumerate(STATUSES, start=2):
        enum.cell(i, 4).value = value
    validations = [
        ("C2:C500", f"={ENUM_SHEET}!$A$2:$A${len(TYPES) + 1}"),
        ("E2:E500", f"={ENUM_SHEET}!$B$2:$B${len(BUSES) + 1}"),
        ("G2:G500", f"={ENUM_SHEET}!$D$2:$D${len(STATUSES) + 1}"),
        ("O2:O500", f"={ENUM_SHEET}!$C$2:$C${len(PHASES) + 1}"),
    ]
    for cell_range, formula in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(cell_range)


def write_workbook(audio: dict[str, Any], workbook_path: Path = WORKBOOK_PATH) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(HEADERS)
    for asset_id in sorted(audio.keys(), key=_sort_key):
        row = audio[asset_id]
        ws.append([
            asset_id,
            row.get("display_name_zh", ""),
            row.get("type", ""),
            row.get("group", ""),
            row.get("bus", ""),
            row.get("path", ""),
            row.get("asset_status", "planned"),
            bool(row.get("loop", False)),
            row.get("volume_db", 0.0),
            row.get("pitch_min", 1.0),
            row.get("pitch_max", 1.0),
            row.get("max_polyphony", 1),
            row.get("priority", 0),
            row.get("suggested_duration_s", ""),
            row.get("implementation_phase", ""),
            row.get("trigger", ""),
            row.get("effect_notes", ""),
            row.get("source_notes", ""),
        ])
    _style_header(ws)
    _set_widths(ws)
    _add_column_guide(wb)
    _add_validations(wb, ws)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def _sort_key(asset_id: str) -> tuple[int, str]:
    prefixes = [
        "music_", "amb_", "stinger_", "ui_", "elixir_", "run_", "reward_",
        "relic_", "meta_", "deploy_", "attack_", "hit_", "death_", "bow_",
        "arrow_", "magic_", "fire_", "tower_", "spell_",
    ]
    for i, prefix in enumerate(prefixes):
        if asset_id.startswith(prefix):
            return (i, asset_id)
    return (999, asset_id)


def init_catalog() -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    missing: list[str] = []
    for base_row in DEFAULT_ROWS:
        row = dict(base_row)
        asset_id = str(row["asset_id"])
        meta = ZH_AUDIO_META.get(asset_id)
        if meta is None:
            missing.append(asset_id)
            continue
        row["display_name_zh"] = meta[0]
        row["effect_notes"] = meta[1]
        row["source_notes"] = row.get("source_notes") or "未入库；待制作/采购音频文件"
        rows.append(row)
    if missing:
        raise AudioConfigError("missing Chinese audio metadata: " + ", ".join(missing))
    return rows_to_audio(rows)


def build_json_from_workbook() -> dict[str, Any]:
    return rows_to_audio(_read_rows())


def check() -> None:
    from_workbook = build_json_from_workbook()
    existing = _load_json(AUDIO_JSON)
    if from_workbook != existing:
        raise AudioConfigError("audio config check failed: AudioConfig.xlsx and audio_assets.json differ")
    print("audio config check ok")


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--init", action="store_true", help="create the initial workbook and JSON from the built-in catalog")
    parser.add_argument("--from-json", action="store_true", help="rebuild AudioConfig.xlsx from audio_assets.json")
    parser.add_argument("--check", action="store_true", help="verify workbook and JSON are identical")
    args = parser.parse_args(argv)

    try:
        if args.check:
            check()
            return 0
        if args.init:
            audio = init_catalog()
            _write_json(AUDIO_JSON, audio)
            write_workbook(audio)
            print(f"wrote {AUDIO_JSON.relative_to(ROOT)}")
            print(f"wrote {WORKBOOK_PATH.relative_to(ROOT)}")
            return 0
        if args.from_json:
            audio = _load_json(AUDIO_JSON)
            write_workbook(audio)
            print(f"wrote {WORKBOOK_PATH.relative_to(ROOT)}")
            return 0
        audio = build_json_from_workbook()
        _write_json(AUDIO_JSON, audio)
        print(f"wrote {AUDIO_JSON.relative_to(ROOT)}")
        return 0
    except AudioConfigError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
